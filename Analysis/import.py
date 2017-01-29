import os
import openpyxl
from openpyxl.styles import PatternFill,Border,Side

rootDir = "data"
processedDir = "processed"
for dataFile in os.listdir(rootDir):
    if dataFile.endswith(".txt"):
        print("processing ", dataFile)
        teamNumber,secondPart = dataFile.split('-')
        matchNumber,extra = secondPart.split('.')
        print("\t","teamNumber:",teamNumber,"matchNumber:",matchNumber)
        fileName = os.path.join(rootDir, dataFile)
        with open(fileName) as f:
            x = f.read().splitlines()
        y=x[0]
        z=y.split(',')
        excelFile = "scouting.xlsx"
        xfile = openpyxl.load_workbook(excelFile)
        sheet = xfile.get_sheet_by_name('Sheet1')
        sheet.append([int(teamNumber),
                      int(matchNumber),
                      int(z[0]),
                      int(z[1]),
                      int(z[2]),
                      int(z[3]),
                      int(z[4]),
                      int(z[5]),
                      int(z[6]),
                      int(z[7]),
                      int(z[8]),
                      int(z[9]),
                      int(z[10]),
                      int(z[11]),
                      int(z[12]),
                      int(z[13]),
                      int(z[14]),
                      int(z[15])])
        lastRow = sheet.max_row
        # print(lastRow)
        #auto
        cellVal = 'A' + str(lastRow)
        sheet[cellVal].fill = PatternFill("solid", fgColor="c5d9f1")
        sheet[cellVal].border = Border(right = Side(border_style='thin', color='FF000000'), 
                                       bottom = Side(border_style='thin', color='FF000000'),
                                       left = Side(border_style='thin', color='FF000000'))
        cellVal = 'B' + str(lastRow)
        sheet[cellVal].fill = PatternFill("solid", fgColor="c5d9f1")
        sheet[cellVal].border = Border(right = Side(border_style='thin', color='FF000000'), 
                                       bottom = Side(border_style='thin', color='FF000000'),
                                       left = Side(border_style='thin', color='FF000000'))

        cellVal = 'C' + str(lastRow)
        sheet[cellVal].fill = PatternFill("solid", fgColor="fde9d9")
        sheet[cellVal].border = Border(right = Side(border_style='thin', color='FF000000'), 
                                       bottom = Side(border_style='thin', color='FF000000'),
                                       left = Side(border_style='medium', color='FF000000'))
        cellVal = 'D' + str(lastRow)
        sheet[cellVal].fill = PatternFill("solid", fgColor="fde9d9")
        sheet[cellVal].border = Border(right = Side(border_style='thin', color='FF000000'), 
                                       bottom = Side(border_style='thin', color='FF000000'),
                                       left = Side(border_style='thin', color='FF000000'))
        cellVal = 'E' + str(lastRow)
        sheet[cellVal].fill = PatternFill("solid", fgColor="fde9d9")
        sheet[cellVal].border = Border(right = Side(border_style='thin', color='FF000000'), 
                                       bottom = Side(border_style='thin', color='FF000000'),
                                       left = Side(border_style='thin', color='FF000000'))
        cellVal = 'F' + str(lastRow)
        sheet[cellVal].fill = PatternFill("solid", fgColor="fde9d9")
        sheet[cellVal].border = Border(right = Side(border_style='thin', color='FF000000'), 
                                       bottom = Side(border_style='thin', color='FF000000'),
                                       left = Side(border_style='thin', color='FF000000'))
        cellVal = 'G' + str(lastRow)
        sheet[cellVal].fill = PatternFill("solid", fgColor="fde9d9")
        sheet[cellVal].border = Border(right = Side(border_style='thin', color='FF000000'), 
                                       bottom = Side(border_style='thin', color='FF000000'),
                                       left = Side(border_style='thin', color='FF000000'))
        cellVal = 'H' + str(lastRow)
        sheet[cellVal].fill = PatternFill("solid", fgColor="fde9d9")
        sheet[cellVal].border = Border(right = Side(border_style='thin', color='FF000000'), 
                                       bottom = Side(border_style='thin', color='FF000000'),
                                       left = Side(border_style='thin', color='FF000000'))
        cellVal = 'I' + str(lastRow)
        sheet[cellVal].fill = PatternFill("solid", fgColor="fde9d9")
        sheet[cellVal].border = Border(right = Side(border_style='thin', color='FF000000'), 
                                       bottom = Side(border_style='thin', color='FF000000'),
                                       left = Side(border_style='thin', color='FF000000'))
        cellVal = 'J' + str(lastRow)
        sheet[cellVal].fill = PatternFill("solid", fgColor="fde9d9")
        sheet[cellVal].border = Border(right = Side(border_style='thin', color='FF000000'), 
                                       bottom = Side(border_style='thin', color='FF000000'),
                                       left = Side(border_style='thin', color='FF000000'))
        cellVal = 'K' + str(lastRow)
        sheet[cellVal].fill = PatternFill("solid", fgColor="fde9d9")
        sheet[cellVal].border = Border(right = Side(border_style='thin', color='FF000000'), 
                                       bottom = Side(border_style='thin', color='FF000000'),
                                       left = Side(border_style='thin', color='FF000000'))
        
        #teleop
        cellVal = 'L' + str(lastRow)
        sheet[cellVal].fill = PatternFill("solid", fgColor="f2dcdb")
        sheet[cellVal].border = Border(right = Side(border_style='thin', color='FF000000'), 
                                       bottom = Side(border_style='thin', color='FF000000'),
                                       left = Side(border_style='medium', color='FF000000'))
        cellVal = 'M' + str(lastRow)
        sheet[cellVal].fill = PatternFill("solid", fgColor="f2dcdb")
        sheet[cellVal].border = Border(right = Side(border_style='thin', color='FF000000'), 
                                       bottom = Side(border_style='thin', color='FF000000'),
                                       left = Side(border_style='thin', color='FF000000'))
        cellVal = 'N' + str(lastRow)
        sheet[cellVal].fill = PatternFill("solid", fgColor="f2dcdb")
        sheet[cellVal].border = Border(right = Side(border_style='thin', color='FF000000'), 
                                       bottom = Side(border_style='thin', color='FF000000'),
                                       left = Side(border_style='thin', color='FF000000'))

        #end game
        cellVal = 'O' + str(lastRow)
        sheet[cellVal].fill = PatternFill("solid", fgColor="ebf1de")
        sheet[cellVal].border = Border(right = Side(border_style='thin', color='FF000000'), 
                                       bottom = Side(border_style='thin', color='FF000000'),
                                       left = Side(border_style='medium', color='FF000000'))
        cellVal = 'P' + str(lastRow)
        sheet[cellVal].fill = PatternFill("solid", fgColor="ebf1de")
        sheet[cellVal].border = Border(right = Side(border_style='thin', color='FF000000'), 
                                       bottom = Side(border_style='thin', color='FF000000'),
                                       left = Side(border_style='thin', color='FF000000'))
        cellVal = 'Q' + str(lastRow)
        sheet[cellVal].fill = PatternFill("solid", fgColor="ebf1de")
        sheet[cellVal].border = Border(right = Side(border_style='thin', color='FF000000'), 
                                       bottom = Side(border_style='thin', color='FF000000'),
                                       left = Side(border_style='thin', color='FF000000'))
        cellVal = 'R' + str(lastRow)
        sheet[cellVal].fill = PatternFill("solid", fgColor="ebf1de")
        sheet[cellVal].border = Border(right = Side(border_style='medium', color='FF000000'), 
                                       bottom = Side(border_style='thin', color='FF000000'),
                                       left = Side(border_style='thin', color='FF000000'))

        xfile.save(excelFile)
        if not os.path.exists(processedDir):
            os.makedirs(processedDir)
        processedName = os.path.join(processedDir, dataFile)
        os.rename(fileName,processedName)